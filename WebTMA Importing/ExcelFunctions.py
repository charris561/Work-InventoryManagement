#This class provides the methods to handle excel documents
#Author: Caleb Harris - UCCS OIT Services Professional
#Date Created: 4/11/2022

from Asset import Asset
import constants
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import GUI

def assignDataFields(asset, cellValue, currentHeaderVal):
    '''
    Function assigns passed in value to field in passed in asset object
    Note: The string compared to currentHeaderVal must be exactly the same as the columns in excel.
        The headers may or may not be present in the excel file. 
    '''
    if (cellValue != None):
        if (currentHeaderVal == 'Tag Number'):
            asset.setAssetTag(cellValue)
        elif (currentHeaderVal == 'Description'):
            asset.setDescription(cellValue)
        elif (currentHeaderVal == 'Building Name'):
            asset.setBuilding(cellValue)
        elif (currentHeaderVal == 'Area #'):
            asset.setAreaNum(cellValue)
        elif (currentHeaderVal == 'Serial #'):
            asset.setSerialNum(cellValue)
        elif (currentHeaderVal == 'Model #'):
            asset.setModelNum(cellValue)
        elif (currentHeaderVal == 'Manufacturer'):
            asset.setManufacturer(cellValue)
        elif (currentHeaderVal == 'Type'):
            asset.setType(cellValue)
        elif (currentHeaderVal == 'Quantity'):
            asset.setQuantity(cellValue)
        elif (currentHeaderVal == 'Warranty Expires'):
            asset.setWarrantyDate(f"{cellValue}")
        elif (currentHeaderVal == 'Assigned To'):
            asset.setUserAssignedTo(cellValue)
        elif (currentHeaderVal == 'Comment'):
            asset.setComment(cellValue)
        elif (currentHeaderVal == 'Facility'):
            asset.setFacility(cellValue)
        elif (currentHeaderVal == 'Subtype'):
            asset.setSubtype(cellValue)
        elif (currentHeaderVal == 'Vendor'):
            asset.setVendor(cellValue)
        else:
            raise Exception(f"Header value: {currentHeaderVal} is unknown.")
    
def getDataFromExcel(incomingAsset_List):
    """
    Function gets data from excel and stores in passed in data structure
    """

    #define workbook
    constants.INVENTORY_SPREADSHEET_FILENAME = GUI.getFilePath()
    inventoryWorkbook = load_workbook(filename = constants.INVENTORY_SPREADSHEET_FILENAME)
    
    #define sheet
    print("\nExcel sheets in this file: \n------------------------------")
    sheetNames = inventoryWorkbook.sheetnames
    for sheet in sheetNames:
        print(sheet)
    print("------------------------------\n")
    constants.INCOMING_DATA_SHEETNAME = GUI.getSheetName()
    incomingDataSheet = inventoryWorkbook[constants.INCOMING_DATA_SHEETNAME]

    #find rows and cols of sheet including header line
    rows = incomingDataSheet.max_row
    cols = incomingDataSheet.max_column

    #store data in object then store object in data structure
    for row in range(rows):
        asset = Asset()
        for col in range(cols):
            colLetter = get_column_letter(col+1)
            currentHeaderVal = incomingDataSheet[f"{colLetter}1"].value
            currentCellVal = incomingDataSheet[f"{colLetter}{row + 1}"].value
            
            if (currentCellVal != currentHeaderVal):

                if (currentCellVal != ''):
                    assignDataFields(asset, currentCellVal, currentHeaderVal)

                if (col == cols - 1):
                    incomingAsset_List.append(asset)